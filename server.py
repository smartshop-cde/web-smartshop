from __future__ import annotations

import json
import mimetypes
import random
import re
import sqlite3
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
DB_PATH = DATA_DIR / "smartshop.sqlite3"
SEED_PATH = DATA_DIR / "seed.json"
DEFAULT_PRODUCT_IMAGE = "assets/logo-smartshop.png"
CANONICAL_DOMAIN = "smartshop.com.py"
WWW_DOMAIN = f"www.{CANONICAL_DOMAIN}"
PRODUCT_CODE_MIN = 10000
PRODUCT_CODE_MAX = 99999
PRODUCT_CODE_RE = re.compile(r"^\d{5}$")


def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def slugify(value: str, fallback: str = "item") -> str:
    value = (value or fallback).strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    return value or fallback


def is_valid_product_code(value: object) -> bool:
    return bool(PRODUCT_CODE_RE.fullmatch(str(value or "").strip()))


def generate_product_code(used_codes: set[str]) -> str:
    for _ in range(1000):
        code = str(random.randint(PRODUCT_CODE_MIN, PRODUCT_CODE_MAX))
        if code not in used_codes:
            used_codes.add(code)
            return code
    raise RuntimeError("No se pudo generar un codigo unico de producto.")


def table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}


def ensure_schema(conn: sqlite3.Connection) -> None:
    columns = table_columns(conn, "products")
    if "code" not in columns:
        conn.execute("ALTER TABLE products ADD COLUMN code TEXT")
    if "brand" not in columns:
        conn.execute("ALTER TABLE products ADD COLUMN brand TEXT DEFAULT ''")
    if "variant" not in columns:
        conn.execute("ALTER TABLE products ADD COLUMN variant TEXT DEFAULT ''")
    assign_missing_product_codes(conn)
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_products_code_unique ON products(code)")


def assign_missing_product_codes(conn: sqlite3.Connection) -> None:
    rows = conn.execute("SELECT id, code FROM products ORDER BY sort_order, name").fetchall()
    used_codes: set[str] = set()
    for row in rows:
        current_code = str(row["code"] or "").strip()
        if is_valid_product_code(current_code) and current_code not in used_codes:
            used_codes.add(current_code)
            continue
        code = generate_product_code(used_codes)
        conn.execute("UPDATE products SET code = ? WHERE id = ?", (code, row["id"]))


def init_db() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS store_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS products (
                id TEXT PRIMARY KEY,
                code TEXT UNIQUE,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                sku TEXT NOT NULL UNIQUE,
                price REAL NOT NULL DEFAULT 0,
                stock INTEGER NOT NULL DEFAULT 0,
                featured INTEGER NOT NULL DEFAULT 0,
                badge TEXT DEFAULT '',
                brand TEXT DEFAULT '',
                variant TEXT DEFAULT '',
                condition TEXT DEFAULT '',
                warranty TEXT DEFAULT '',
                delivery TEXT DEFAULT '',
                description TEXT DEFAULT '',
                details TEXT NOT NULL DEFAULT '[]',
                image TEXT DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS sellers (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                role TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                schedule TEXT DEFAULT '',
                message TEXT DEFAULT '',
                image TEXT DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0
            );
            """
        )
        ensure_schema(conn)
        count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        if count == 0 and SEED_PATH.exists():
            save_catalog(conn, json.loads(SEED_PATH.read_text(encoding="utf-8")))
            ensure_schema(conn)


def get_catalog() -> dict:
    with connect() as conn:
        store_rows = conn.execute("SELECT key, value FROM store_settings").fetchall()
        store = {row["key"]: json.loads(row["value"]) for row in store_rows}
        products = [
            {
                **dict(row),
                "code": str(row["code"] or ""),
                "featured": bool(row["featured"]),
                "details": json.loads(row["details"] or "[]"),
            }
            for row in conn.execute("SELECT * FROM products ORDER BY sort_order, name").fetchall()
        ]
        sellers = [
            dict(row)
            for row in conn.execute("SELECT * FROM sellers ORDER BY sort_order, name").fetchall()
        ]
    return {"store": store, "products": products, "sellers": sellers}


def save_catalog(conn: sqlite3.Connection, catalog: dict) -> None:
    store = catalog.get("store") or {}
    products = catalog.get("products") or []
    sellers = catalog.get("sellers") or []
    existing_products = conn.execute("SELECT id, sku, code FROM products").fetchall()
    existing_codes_by_id = {
        row["id"]: str(row["code"]) for row in existing_products if is_valid_product_code(row["code"])
    }
    existing_codes_by_sku = {
        row["sku"]: str(row["code"]) for row in existing_products if is_valid_product_code(row["code"])
    }
    used_codes: set[str] = set()

    conn.execute("DELETE FROM store_settings")
    conn.execute("DELETE FROM products")
    conn.execute("DELETE FROM sellers")

    for key, value in store.items():
        conn.execute(
            "INSERT INTO store_settings (key, value) VALUES (?, ?)",
            (key, json.dumps(value, ensure_ascii=False)),
        )

    for index, product in enumerate(products):
        product_id = product.get("id") or slugify(product.get("sku") or product.get("name"), "product")
        incoming_code = str(product.get("code") or "").strip()
        existing_code = existing_codes_by_id.get(product_id) or existing_codes_by_sku.get(product.get("sku") or "")
        if is_valid_product_code(incoming_code) and incoming_code not in used_codes:
            code = incoming_code
            used_codes.add(code)
        elif existing_code and existing_code not in used_codes:
            code = existing_code
            used_codes.add(code)
        else:
            code = generate_product_code(used_codes)
        sku = product.get("sku") or f"SKU-{code}"
        details = product.get("details") if isinstance(product.get("details"), list) else []
        conn.execute(
            """
            INSERT INTO products (
                id, code, name, category, sku, price, stock, featured, badge, brand, variant, condition,
                warranty, delivery, description, details, image, sort_order
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                product_id,
                code,
                product.get("name") or "Producto",
                product.get("category") or "General",
                sku,
                float(product.get("price") or 0),
                int(product.get("stock") or 0),
                1 if product.get("featured") else 0,
                product.get("badge") or "",
                product.get("brand") or "",
                product.get("variant") or "",
                product.get("condition") or "",
                product.get("warranty") or "",
                product.get("delivery") or "",
                product.get("description") or "",
                json.dumps(details, ensure_ascii=False),
                product.get("image") or DEFAULT_PRODUCT_IMAGE,
                index,
            ),
        )

    for index, seller in enumerate(sellers):
        seller_id = seller.get("id") or slugify(seller.get("name"), f"seller-{index + 1}")
        conn.execute(
            """
            INSERT INTO sellers (id, name, role, phone, schedule, message, image, sort_order)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                seller_id,
                seller.get("name") or "Vendedor",
                seller.get("role") or "",
                seller.get("phone") or "",
                seller.get("schedule") or "",
                seller.get("message") or "",
                seller.get("image") or DEFAULT_PRODUCT_IMAGE,
                index,
            ),
        )


def verify_pin(headers, catalog: dict | None = None) -> bool:
    catalog = catalog or get_catalog()
    expected_pin = str((catalog.get("store") or {}).get("adminPin") or "")
    sent_pin = headers.get("X-Admin-Pin") or ""
    return bool(expected_pin and sent_pin == expected_pin)


def export_products_excel() -> bytes:
    catalog = get_catalog()
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    headers = [
        "accion",
        "code",
        "sku",
        "name",
        "category",
        "price",
        "stock",
        "featured",
        "badge",
        "brand",
        "variant",
        "condition",
        "warranty",
        "delivery",
        "description",
        "details",
    ]
    ws.append(headers)
    for product in catalog["products"]:
        ws.append(
            [
                "",
                product.get("code"),
                product.get("sku"),
                product.get("name"),
                product.get("category"),
                product.get("price"),
                product.get("stock"),
                "SI" if product.get("featured") else "NO",
                product.get("badge"),
                product.get("brand"),
                product.get("variant"),
                product.get("condition"),
                product.get("warranty"),
                product.get("delivery"),
                product.get("description"),
                "\n".join(product.get("details") or []),
            ]
        )
    for cell in ws[1]:
        cell.style = "Headline 3"
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def import_products_excel(body: bytes) -> dict:
    wb = load_workbook(BytesIO(body), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required = {"name", "category", "price", "stock"}
    missing = sorted(required - set(header_index))
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

    catalog = get_catalog()
    products_by_code = {
        str(p.get("code")): p for p in catalog["products"] if is_valid_product_code(p.get("code"))
    }
    products_by_sku = {str(p.get("sku")): p for p in catalog["products"] if p.get("sku")}
    imported = 0
    deleted = 0

    def cell(row, name, default=""):
        idx = header_index.get(name)
        if idx is None:
            return default
        value = row[idx].value
        return default if value is None else value

    for row in ws.iter_rows(min_row=2):
        code = str(cell(row, "code", "")).strip()
        sku = str(cell(row, "sku", "")).strip()
        action = str(cell(row, "accion", "")).strip().lower()
        has_visible_data = any(
            str(cell(row, name, "") or "").strip()
            for name in {"name", "category", "price", "stock", "description", "brand", "variant"}
            if name in header_index
        )
        if not code and not sku and not has_visible_data:
            continue
        current = products_by_code.get(code) if is_valid_product_code(code) else None
        if current is None and sku:
            current = products_by_sku.get(sku)
        if action in {"eliminar", "delete", "quitar", "borrar"}:
            if current and current in catalog["products"]:
                catalog["products"].remove(current)
                deleted += 1
            continue

        current_product = current or {}
        details_value = str(cell(row, "details", "") or "")
        product_data = {
            **current_product,
            "id": current_product.get("id") or slugify(sku or code or cell(row, "name", "product"), "product"),
            "code": code if is_valid_product_code(code) else current_product.get("code"),
            "sku": sku or current_product.get("sku") or "",
            "name": str(cell(row, "name", current_product.get("name") or "Producto")).strip(),
            "category": str(cell(row, "category", current_product.get("category") or "General")).strip(),
            "price": float(cell(row, "price", current_product.get("price") or 0) or 0),
            "stock": int(cell(row, "stock", current_product.get("stock") or 0) or 0),
            "featured": str(cell(row, "featured", current_product.get("featured") or "NO")).strip().lower()
            in {"si", "sí", "true", "1", "yes"},
            "badge": str(cell(row, "badge", current_product.get("badge") or "")).strip(),
            "brand": str(cell(row, "brand", current_product.get("brand") or "")).strip(),
            "variant": str(cell(row, "variant", current_product.get("variant") or "")).strip(),
            "condition": str(cell(row, "condition", current_product.get("condition") or "Nuevo")).strip(),
            "warranty": str(cell(row, "warranty", current_product.get("warranty") or "Garantia de tienda")).strip(),
            "delivery": str(cell(row, "delivery", current_product.get("delivery") or "Retiro en tienda o envio coordinado")).strip(),
            "description": str(cell(row, "description", current_product.get("description") or "")).strip(),
            "details": [line.strip() for line in details_value.splitlines() if line.strip()]
            or current_product.get("details")
            or [],
            "image": current_product.get("image") or DEFAULT_PRODUCT_IMAGE,
        }
        if current in catalog["products"]:
            catalog["products"][catalog["products"].index(current)] = product_data
        else:
            catalog["products"].append(product_data)
        imported += 1

    with connect() as conn:
        save_catalog(conn, catalog)
    return {"imported": imported, "deleted": deleted, "total": len(catalog["products"])}


class SmartShopHandler(SimpleHTTPRequestHandler):
    server_version = "SmartShopHTTP/1.0"

    def redirect_www_to_canonical(self) -> bool:
        host = (self.headers.get("Host") or "").split(":", 1)[0].lower()
        if host != WWW_DOMAIN:
            return False
        self.send_response(HTTPStatus.MOVED_PERMANENTLY)
        self.send_header("Location", f"https://{CANONICAL_DOMAIN}{self.path}")
        self.send_header("Cache-Control", "public, max-age=3600")
        self.end_headers()
        return True

    def translate_path(self, path: str) -> str:
        path = unquote(path.split("?", 1)[0].split("#", 1)[0]).lstrip("/")
        resolved = (ROOT / path).resolve()
        if resolved != ROOT and ROOT not in resolved.parents:
            return str(ROOT / "index.html")
        return str(resolved)

    def do_GET(self) -> None:
        if self.redirect_www_to_canonical():
            return
        if self.path == "/api/catalog":
            return self.send_json(get_catalog())
        if self.path == "/api/products/export-excel":
            if not verify_pin(self.headers):
                return self.send_json({"error": "No autorizado"}, HTTPStatus.UNAUTHORIZED)
            body = export_products_excel()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="smartshop-productos.xlsx"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        return super().do_GET()

    def do_POST(self) -> None:
        if self.redirect_www_to_canonical():
            return
        if self.path == "/api/login":
            body = self.read_json()
            pin = str((body or {}).get("pin") or "")
            if verify_pin({"X-Admin-Pin": pin}):
                return self.send_json({"ok": True})
            return self.send_json({"ok": False, "error": "PIN incorrecto"}, HTTPStatus.UNAUTHORIZED)
        if self.path == "/api/products/import-excel":
            if not verify_pin(self.headers):
                return self.send_json({"error": "No autorizado"}, HTTPStatus.UNAUTHORIZED)
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length)
            try:
                result = import_products_excel(body)
            except Exception as exc:  # noqa: BLE001
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return self.send_json(result)
        return self.send_json({"error": "Ruta no encontrada"}, HTTPStatus.NOT_FOUND)

    def do_PUT(self) -> None:
        if self.redirect_www_to_canonical():
            return
        if self.path != "/api/catalog":
            return self.send_json({"error": "Ruta no encontrada"}, HTTPStatus.NOT_FOUND)
        if not verify_pin(self.headers):
            return self.send_json({"error": "No autorizado"}, HTTPStatus.UNAUTHORIZED)
        catalog = self.read_json()
        with connect() as conn:
            save_catalog(conn, catalog or {})
        return self.send_json({"ok": True})

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length") or 0)
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def guess_type(self, path: str) -> str:
        if path.endswith(".webp"):
            return "image/webp"
        return mimetypes.guess_type(path)[0] or "application/octet-stream"


def main() -> None:
    init_db()
    server = ThreadingHTTPServer(("127.0.0.1", 8000), SmartShopHandler)
    print("SmartShop disponible en http://127.0.0.1:8000")
    server.serve_forever()


if __name__ == "__main__":
    main()
